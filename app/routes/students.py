from __future__ import annotations

from typing import List, Optional, Dict, Any
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_
from openpyxl import Workbook, load_workbook

from .. import models, schemas
from ..database import get_db
from ..dependencies import get_active_member, get_admin_user
from ..utils import generate_code
from datetime import datetime, timedelta

router = APIRouter(prefix="/api/students", tags=["students"])
DEFAULT_SUBJECTS = ["语文", "数学", "英语", "物理", "化学", "生物", "历史", "政治", "地理"]
ALL_SUBJECTS = ["总分"] + DEFAULT_SUBJECTS
DEFAULT_RANGE_CONFIG = [
    {"key": "fail", "name": "不及格", "min": 0, "max": 60},
    {"key": "pass", "name": "及格", "min": 60, "max": 80},
    {"key": "good", "name": "良好", "min": 80, "max": 90},
    {"key": "excellent", "name": "优秀", "min": 90, "max": 100},
]



def _serialize_student(student: models.Student) -> schemas.Student:
    scores = student.scores or []
    total = sum(item.get("score", 0) for item in scores)
    average = total / len(scores) if scores else 0.0
    return schemas.Student(
        id=student.id,
        name=student.name,
        student_no=student.student_no,
        class_name=student.class_name,
        grade_name=student.grade_name,
        exam_name=student.exam_name,
        gender=student.gender,
        notes=student.notes,
        scores=[schemas.ScoreItem(**item) for item in scores],
        class_rank=student.class_rank,
        grade_rank=student.grade_rank,
        total_score=round(total, 2),
        average_score=round(average, 2),
        created_at=student.created_at,
        updated_at=student.updated_at,
    )


@router.get("/", response_model=List[schemas.Student])
def list_students(
    keyword: Optional[str] = Query(None, description="按姓名/学号模糊搜索"),
    class_name: Optional[str] = None,
    exam_name: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.Member = Depends(get_active_member),
):
    query = db.query(models.Student)
    if keyword:
        like = f"%{keyword}%"
        query = query.filter(
            or_(
                models.Student.name.ilike(like),
                models.Student.student_no.ilike(like),
            )
        )
    if class_name:
        query = query.filter(models.Student.class_name == class_name)
    if exam_name:
        query = query.filter(models.Student.exam_name == exam_name)
    students = query.order_by(models.Student.updated_at.desc()).all()
    return [_serialize_student(s) for s in students]


@router.get("/summary", response_model=schemas.StudentSummary)
def summary(
    keyword: Optional[str] = Query(None, description="按姓名/学号模糊搜索"),
    class_name: Optional[str] = None,
    exam_name: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """获取学生统计摘要，支持按班级、考试筛选"""
    query = db.query(models.Student)

    # 应用筛选条件
    if keyword:
        like = f"%{keyword}%"
        query = query.filter(
            or_(
                models.Student.name.ilike(like),
                models.Student.student_no.ilike(like),
            )
        )
    if class_name:
        query = query.filter(models.Student.class_name == class_name)
    if exam_name:
        query = query.filter(models.Student.exam_name == exam_name)

    students = query.all()

    if not students:
        return schemas.StudentSummary(
            total_students=0,
            average_total=0.0,
            highest_total=None,
            highest_student=None,
            pass_rate=0.0,
            good_rate=0.0,
            excellent_rate=0.0,
        )

    # 获取总分的成绩区间配置
    total_score_range = _ensure_subject_range(db, "总分")
    range_config = total_score_range.config

    # 提取各等级的最小分数线
    pass_min = next((r.get("min", 60) for r in range_config if r.get("key") == "pass"), 60)
    good_min = next((r.get("min", 80) for r in range_config if r.get("key") == "good"), 80)
    excellent_min = next((r.get("min", 90) for r in range_config if r.get("key") == "excellent"), 90)

    totals = []
    highest_total = None
    highest_name = None
    pass_count = 0
    good_count = 0
    excellent_count = 0
    for student in students:
        scores = student.scores or []
        total = sum(item.get("score", 0) for item in scores)
        totals.append(total)
        if highest_total is None or total > highest_total:
            highest_total = total
            highest_name = student.name
        avg = total / len(scores) if scores else 0
        # 使用配置的分数线
        if avg >= excellent_min:
            excellent_count += 1
        if avg >= good_min:
            good_count += 1
        if avg >= pass_min:
            pass_count += 1
    total_students = len(students)
    average_total = sum(totals) / total_students if totals else 0
    return schemas.StudentSummary(
        total_students=total_students,
        average_total=round(average_total, 2),
        highest_total=round(highest_total, 2) if highest_total is not None else None,
        highest_student=highest_name,
        pass_rate=round(pass_count / total_students * 100, 2),
        good_rate=round(good_count / total_students * 100, 2),
        excellent_rate=round(excellent_count / total_students * 100, 2),
    )


@router.get("/class-stats", response_model=List[schemas.ClassStats])
def get_class_stats(db: Session = Depends(get_db)):
    students = db.query(models.Student).all()
    if not students:
        return []

    # 获取总分的成绩区间配置
    total_score_range = _ensure_subject_range(db, "总分")
    range_config = total_score_range.config

    # 提取各等级的最小分数线
    pass_min = next((r.get("min", 60) for r in range_config if r.get("key") == "pass"), 60)
    good_min = next((r.get("min", 80) for r in range_config if r.get("key") == "good"), 80)
    excellent_min = next((r.get("min", 90) for r in range_config if r.get("key") == "excellent"), 90)

    class_data = {}
    for student in students:
        class_name = student.class_name or "未分配班级"
        if class_name not in class_data:
            class_data[class_name] = {
                "records": [],
                "grade_name": student.grade_name,
                "unique_students": set(),
                "exams": set(),
            }
        class_data[class_name]["records"].append(student)
        class_data[class_name]["unique_students"].add(student.student_no)
        if student.exam_name:
            class_data[class_name]["exams"].add(student.exam_name)

    result = []
    for class_name, data in class_data.items():
        records = data["records"]
        total_students = len(data["unique_students"])
        exam_count = len(data["exams"]) if data["exams"] else 1

        male_students = set()
        female_students = set()
        for record in records:
            if record.gender == "男":
                male_students.add(record.student_no)
            elif record.gender == "女":
                female_students.add(record.student_no)

        male_count = len(male_students)
        female_count = len(female_students)
        male_ratio = round(male_count / total_students * 100, 2) if total_students > 0 else 0.0
        female_ratio = round(female_count / total_students * 100, 2) if total_students > 0 else 0.0

        total_avg = 0
        pass_count = 0
        good_count = 0
        excellent_count = 0
        valid_records = 0

        for record in records:
            scores = record.scores or []
            if scores:
                total = sum(item.get("score", 0) for item in scores)
                avg = total / len(scores)
                total_avg += avg
                valid_records += 1

                # 使用配置的分数线
                if avg >= excellent_min:
                    excellent_count += 1
                if avg >= good_min:
                    good_count += 1
                if avg >= pass_min:
                    pass_count += 1

        average_score = round(total_avg / valid_records, 2) if valid_records > 0 else 0.0
        pass_rate = round(pass_count / valid_records * 100, 2) if valid_records > 0 else 0.0
        good_rate = round(good_count / valid_records * 100, 2) if valid_records > 0 else 0.0
        excellent_rate = round(excellent_count / valid_records * 100, 2) if valid_records > 0 else 0.0

        result.append(schemas.ClassStats(
            class_name=class_name,
            grade_name=data["grade_name"],
            total_students=total_students,
            exam_count=exam_count,
            male_count=male_count,
            female_count=female_count,
            male_ratio=male_ratio,
            female_ratio=female_ratio,
            average_score=average_score,
            pass_rate=pass_rate,
            good_rate=good_rate,
            excellent_rate=excellent_rate,
        ))

    return sorted(result, key=lambda x: (x.grade_name or "", x.class_name))


@router.put("/class/{class_name}/rename")
def rename_class(
    class_name: str,
    new_class_name: str = Query(..., description="新班级名称"),
    new_grade_name: Optional[str] = Query(None, description="新年级名称"),
    db: Session = Depends(get_db)
):
    """重命名班级（批量更新所有该班级的学生记录）"""
    students = db.query(models.Student).filter(models.Student.class_name == class_name).all()

    if not students:
        raise HTTPException(status_code=404, detail="未找到该班级的学生")

    for student in students:
        student.class_name = new_class_name
        if new_grade_name is not None:
            student.grade_name = new_grade_name
        db.add(student)

    db.commit()
    return {"message": f"已更新 {len(students)} 条记录", "updated_count": len(students)}


@router.delete("/class/{class_name}")
def delete_class(class_name: str, db: Session = Depends(get_db)):
    """删除班级（删除该班级的所有学生记录）"""
    students = db.query(models.Student).filter(models.Student.class_name == class_name).all()

    if not students:
        raise HTTPException(status_code=404, detail="未找到该班级的学生")

    count = len(students)
    for student in students:
        db.delete(student)

    db.commit()
    return {"message": f"已删除班级 {class_name} 及其 {count} 条学生记录", "deleted_count": count}


@router.post("/class/{class_name}/share")
def create_class_share_code(
    class_name: str,
    valid_days: int = Query(30, ge=1, le=365, description="有效天数"),
    exam_name: str = Query(None, description="考试名称，为空则查询所有考试"),
    db: Session = Depends(get_db)
):
    """为班级生成分享查询码"""
    query = db.query(models.Student).filter(models.Student.class_name == class_name)
    if exam_name:
        query = query.filter(models.Student.exam_name == exam_name)

    students = query.first()
    if not students:
        raise HTTPException(status_code=404, detail="未找到该班级的学生")

    code = generate_code()
    expires_at = datetime.utcnow() + timedelta(days=valid_days)

    description = f"班级 {class_name}"
    if exam_name:
        description += f" - {exam_name}"
    description += " 的成绩查询"

    query_code = models.QueryCode(
        code=code,
        expires_at=expires_at,
        description=description,
        is_active=True,
    )
    db.add(query_code)
    db.commit()
    db.refresh(query_code)

    return {
        "code": code,
        "class_name": class_name,
        "exam_name": exam_name,
        "expires_at": expires_at,
        "query_url": f"/query?code={code}"
    }


@router.get("/query-by-code")
def query_student_by_code(
    code: str = Query(..., description="查询码"),
    keyword: str = Query(..., description="学生姓名或学号"),
    db: Session = Depends(get_db)
):
    """使用查询码查询学生成绩"""
    query_code = db.query(models.QueryCode).filter(
        models.QueryCode.code == code,
        models.QueryCode.is_active == True
    ).first()

    if not query_code:
        raise HTTPException(status_code=404, detail="查询码无效")

    if query_code.expires_at <= datetime.utcnow():
        query_code.is_active = False
        db.commit()
        raise HTTPException(status_code=400, detail="查询码已过期")

    query_code.last_used_at = datetime.utcnow()
    db.commit()

    # 精确匹配查询：姓名或学号必须完全一致
    query = db.query(models.Student).filter(
        or_(
            models.Student.name == keyword,
            models.Student.student_no == keyword
        )
    )

    students = query.order_by(models.Student.created_at.desc()).all()

    if not students:
        raise HTTPException(status_code=404, detail="未找到匹配的学生，请检查姓名或学号是否正确")

    return [_serialize_student(s) for s in students]


@router.get("/ranges", response_model=Dict[str, List[schemas.RangeSegment]])
def get_all_ranges(db: Session = Depends(get_db)):
    configs: Dict[str, List[schemas.RangeSegment]] = {}
    for subject in ALL_SUBJECTS:
        record = _ensure_subject_range(db, subject)
        configs[subject] = [schemas.RangeSegment(**segment) for segment in record.config]
    return configs


@router.put("/ranges/{subject}", response_model=List[schemas.RangeSegment])
def update_range(subject: str, payload: List[schemas.RangeSegment], db: Session = Depends(get_db)):
    if subject not in ALL_SUBJECTS:
        raise HTTPException(status_code=400, detail="科目不在允许范围内")
    if not payload:
        raise HTTPException(status_code=400, detail="请选择至少一个区间")
    validated = []
    for segment in payload:
        if segment.max is not None and segment.max <= segment.min:
            raise HTTPException(status_code=400, detail=f"{segment.name} 的最大值需大于最小值")
        validated.append(segment.model_dump())
    record = _ensure_subject_range(db, subject)
    record.config = validated
    db.add(record)
    db.commit()
    db.refresh(record)
    return [schemas.RangeSegment(**segment) for segment in record.config]


@router.get("/student-trend/{student_no}", response_model=schemas.StudentTrend)
def get_student_trend(student_no: str, db: Session = Depends(get_db)):
    """获取学生的历次考试趋势数据"""
    from datetime import datetime

    records = db.query(models.Student).filter(models.Student.student_no == student_no).order_by(models.Student.created_at).all()

    if not records:
        raise HTTPException(status_code=404, detail="未找到该学生的考试记录")

    first_record = records[0]
    exams = []

    for record in records:
        scores = record.scores or []
        total = sum(item.get("score", 0) for item in scores)
        average = total / len(scores) if scores else 0.0

        subject_scores = {item.get("subject"): item.get("score", 0) for item in scores}

        exams.append(schemas.ExamTrendPoint(
            exam_name=record.exam_name or "未命名考试",
            exam_date=record.created_at,
            total_score=round(total, 2),
            average_score=round(average, 2),
            subject_scores=subject_scores
        ))

    return schemas.StudentTrend(
        student_no=student_no,
        name=first_record.name,
        class_name=first_record.class_name,
        exams=exams
    )


@router.get("/class-trend/{class_name}", response_model=schemas.ClassTrend)
def get_class_trend(class_name: str, db: Session = Depends(get_db)):
    """获取班级的历次考试趋势数据"""
    from datetime import datetime

    records = db.query(models.Student).filter(models.Student.class_name == class_name).all()

    if not records:
        raise HTTPException(status_code=404, detail="未找到该班级的考试记录")

    # 获取总分的成绩区间配置
    total_score_range = _ensure_subject_range(db, "总分")
    range_config = total_score_range.config

    # 提取各等级的最小分数线
    pass_min = next((r.get("min", 60) for r in range_config if r.get("key") == "pass"), 60)
    good_min = next((r.get("min", 80) for r in range_config if r.get("key") == "good"), 80)
    excellent_min = next((r.get("min", 90) for r in range_config if r.get("key") == "excellent"), 90)

    exam_data = {}
    for record in records:
        exam_name = record.exam_name or "未命名考试"
        if exam_name not in exam_data:
            exam_data[exam_name] = {
                "records": [],
                "exam_date": record.created_at,
                "grade_name": record.grade_name
            }
        exam_data[exam_name]["records"].append(record)

    trend_data = []
    for exam_name, data in exam_data.items():
        exam_records = data["records"]
        total_avg = 0
        pass_count = 0
        good_count = 0
        excellent_count = 0
        valid_count = 0

        for record in exam_records:
            scores = record.scores or []
            if scores:
                total = sum(item.get("score", 0) for item in scores)
                avg = total / len(scores)
                total_avg += avg
                valid_count += 1

                # 使用配置的分数线
                if avg >= excellent_min:
                    excellent_count += 1
                if avg >= good_min:
                    good_count += 1
                if avg >= pass_min:
                    pass_count += 1

        if valid_count > 0:
            trend_data.append(schemas.ClassTrendPoint(
                exam_name=exam_name,
                exam_date=data["exam_date"],
                average_score=round(total_avg / valid_count, 2),
                pass_rate=round(pass_count / valid_count * 100, 2),
                good_rate=round(good_count / valid_count * 100, 2),
                excellent_rate=round(excellent_count / valid_count * 100, 2),
                student_count=valid_count
            ))

    trend_data.sort(key=lambda x: x.exam_date if x.exam_date else datetime.min)

    grade_name = exam_data[list(exam_data.keys())[0]]["grade_name"] if exam_data else None

    return schemas.ClassTrend(
        class_name=class_name,
        grade_name=grade_name,
        trend_data=trend_data
    )


@router.get("/progress-analysis", response_model=schemas.ProgressAnalysis)
def get_progress_analysis(
    class_name: Optional[str] = Query(None, description="班级名称,为空则分析所有班级"),
    limit: int = Query(10, ge=1, le=50, description="返回TOP N学生"),
    db: Session = Depends(get_db)
):
    """获取进步/退步学生排行榜"""
    query = db.query(models.Student)
    if class_name:
        query = query.filter(models.Student.class_name == class_name)

    all_records = query.all()

    student_data = {}
    for record in all_records:
        student_no = record.student_no
        if student_no not in student_data:
            student_data[student_no] = {
                "name": record.name,
                "class_name": record.class_name,
                "exams": []
            }

        scores = record.scores or []
        if scores:
            total = sum(item.get("score", 0) for item in scores)
            avg = total / len(scores)
            student_data[student_no]["exams"].append({
                "exam_name": record.exam_name or "未命名考试",
                "score": avg,
                "created_at": record.created_at
            })

    progress_list = []
    for student_no, data in student_data.items():
        exams = sorted(data["exams"], key=lambda x: x["created_at"])
        if len(exams) >= 2:
            first_exam = exams[0]
            latest_exam = exams[-1]
            progress = latest_exam["score"] - first_exam["score"]
            progress_rate = (progress / first_exam["score"] * 100) if first_exam["score"] > 0 else 0

            progress_list.append(schemas.ProgressStudent(
                student_no=student_no,
                name=data["name"],
                class_name=data["class_name"],
                first_exam=first_exam["exam_name"],
                first_score=round(first_exam["score"], 2),
                latest_exam=latest_exam["exam_name"],
                latest_score=round(latest_exam["score"], 2),
                progress=round(progress, 2),
                progress_rate=round(progress_rate, 2)
            ))

    progress_list.sort(key=lambda x: x.progress, reverse=True)
    top_progress = progress_list[:limit]

    progress_list.sort(key=lambda x: x.progress)
    top_regress = progress_list[:limit]

    return schemas.ProgressAnalysis(
        top_progress=top_progress,
        top_regress=top_regress
    )


@router.get("/critical-students", response_model=schemas.CriticalStudents)
def get_critical_students(
    class_name: Optional[str] = Query(None, description="班级名称"),
    exam_name: Optional[str] = Query(None, description="考试名称"),
    threshold: float = Query(5.0, ge=0, le=20, description="临界阈值(分数)"),
    db: Session = Depends(get_db)
):
    """识别临界学生(接近及格/良好/优秀线的学生)"""
    query = db.query(models.Student)
    if class_name:
        query = query.filter(models.Student.class_name == class_name)
    if exam_name:
        query = query.filter(models.Student.exam_name == exam_name)

    records = query.all()

    near_pass = []
    near_good = []
    near_excellent = []

    for record in records:
        scores = record.scores or []
        if not scores:
            continue

        total = sum(item.get("score", 0) for item in scores)
        avg = total / len(scores)

        if 60 - threshold <= avg < 60:
            near_pass.append(schemas.CriticalStudent(
                student_no=record.student_no,
                name=record.name,
                class_name=record.class_name,
                exam_name=record.exam_name,
                average_score=round(avg, 2),
                distance_to_line=round(60 - avg, 2),
                category="接近及格"
            ))
        elif 80 - threshold <= avg < 80:
            near_good.append(schemas.CriticalStudent(
                student_no=record.student_no,
                name=record.name,
                class_name=record.class_name,
                exam_name=record.exam_name,
                average_score=round(avg, 2),
                distance_to_line=round(80 - avg, 2),
                category="接近良好"
            ))
        elif 90 - threshold <= avg < 90:
            near_excellent.append(schemas.CriticalStudent(
                student_no=record.student_no,
                name=record.name,
                class_name=record.class_name,
                exam_name=record.exam_name,
                average_score=round(avg, 2),
                distance_to_line=round(90 - avg, 2),
                category="接近优秀"
            ))

    near_pass.sort(key=lambda x: x.distance_to_line)
    near_good.sort(key=lambda x: x.distance_to_line)
    near_excellent.sort(key=lambda x: x.distance_to_line)

    return schemas.CriticalStudents(
        near_pass=near_pass,
        near_good=near_good,
        near_excellent=near_excellent
    )


@router.get("/class-comparison", response_model=List[schemas.ClassComparison])
def get_class_comparison(
    grade_name: Optional[str] = Query(None, description="年级名称,为空则对比所有年级"),
    exam_name: Optional[str] = Query(None, description="考试名称,为空则使用最新考试"),
    db: Session = Depends(get_db)
):
    """班级对比分析"""
    query = db.query(models.Student)
    if grade_name:
        query = query.filter(models.Student.grade_name == grade_name)
    if exam_name:
        query = query.filter(models.Student.exam_name == exam_name)

    records = query.all()

    # 获取总分的成绩区间配置
    total_score_range = _ensure_subject_range(db, "总分")
    range_config = total_score_range.config

    # 提取各等级的最小分数线
    pass_min = next((r.get("min", 60) for r in range_config if r.get("key") == "pass"), 60)
    good_min = next((r.get("min", 80) for r in range_config if r.get("key") == "good"), 80)
    excellent_min = next((r.get("min", 90) for r in range_config if r.get("key") == "excellent"), 90)

    class_data = {}
    for record in records:
        class_name = record.class_name or "未分配班级"
        if class_name not in class_data:
            class_data[class_name] = {
                "records": [],
                "grade_name": record.grade_name
            }
        class_data[class_name]["records"].append(record)

    result = []
    for class_name, data in class_data.items():
        class_records = data["records"]
        total_avg = 0
        pass_count = 0
        good_count = 0
        excellent_count = 0
        valid_count = 0

        for record in class_records:
            scores = record.scores or []
            if scores:
                total = sum(item.get("score", 0) for item in scores)
                avg = total / len(scores)
                total_avg += avg
                valid_count += 1

                # 使用配置的分数线
                if avg >= excellent_min:
                    excellent_count += 1
                if avg >= good_min:
                    good_count += 1
                if avg >= pass_min:
                    pass_count += 1

        if valid_count > 0:
            result.append(schemas.ClassComparison(
                class_name=class_name,
                grade_name=data["grade_name"],
                student_count=valid_count,
                average_score=round(total_avg / valid_count, 2),
                pass_rate=round(pass_count / valid_count * 100, 2),
                good_rate=round(good_count / valid_count * 100, 2),
                excellent_rate=round(excellent_count / valid_count * 100, 2)
            ))

    result.sort(key=lambda x: x.average_score, reverse=True)
    return result


@router.get("/class-subject-stats", response_model=schemas.ClassSubjectStats)
def get_class_subject_stats(
    class_name: str = Query(..., description="班级名称"),
    exam_name: str = Query(..., description="考试名称"),
    db: Session = Depends(get_db)
):
    """获取班级各科目统计数据（用于计算班级平均分等）"""
    records = db.query(models.Student).filter(
        models.Student.class_name == class_name,
        models.Student.exam_name == exam_name
    ).all()

    if not records:
        raise HTTPException(status_code=404, detail="未找到该班级的考试记录")

    # 统计各科目的成绩
    subject_data = {}
    for record in records:
        scores = record.scores or []
        for score_item in scores:
            subject = score_item.get("subject")
            score = score_item.get("score", 0)
            if subject not in subject_data:
                subject_data[subject] = []
            subject_data[subject].append(score)

    # 计算各科目的统计信息
    subject_stats = []
    for subject, scores in subject_data.items():
        if scores:
            subject_stats.append(schemas.SubjectStat(
                subject=subject,
                average=round(sum(scores) / len(scores), 2),
                max_score=round(max(scores), 2),
                min_score=round(min(scores), 2)
            ))

    return schemas.ClassSubjectStats(
        class_name=class_name,
        exam_name=exam_name,
        student_count=len(records),
        subject_stats=subject_stats
    )


@router.get("/template")
def download_template():
    wb = Workbook()
    ws = wb.active
    headers = ["姓名", "学号", "班级", "年级", "考试名称", "考试日期", "备注"] + DEFAULT_SUBJECTS
    ws.append(headers)
    ws.append(["张三", "2024001", "高一(1)班", "高一", "期中考试", "2024-01-15", "可填写备注"] + ["" for _ in DEFAULT_SUBJECTS])
    return _workbook_response(wb, "grade_template.xlsx")


@router.get("/rank-template")
def download_rank_template():
    """下载排名导入模板"""
    wb = Workbook()
    ws = wb.active
    headers = ["姓名", "学号", "考试名称", "年级排名", "班级排名"]
    ws.append(headers)
    ws.append(["张三", "2024001", "期中考试", "15", "3"])
    ws.append(["李四", "2024002", "期中考试", "28", "5"])
    return _workbook_response(wb, "rank_import_template.xlsx")


@router.get("/{student_id}", response_model=schemas.Student)
def get_student(student_id: int, db: Session = Depends(get_db)):
    student = db.get(models.Student, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")
    return _serialize_student(student)


@router.post("/", response_model=schemas.Student, status_code=201)
def create_student(
    payload: schemas.StudentCreate,
    db: Session = Depends(get_db),
    current_user: models.Member = Depends(get_active_member),
):
    if db.query(models.Student).filter_by(student_no=payload.student_no).first():
        raise HTTPException(status_code=400, detail="学号已存在")
    student = models.Student(
        name=payload.name,
        student_no=payload.student_no,
        class_name=payload.class_name,
        grade_name=payload.grade_name,
        exam_name=payload.exam_name,
        gender=payload.gender,
        notes=payload.notes,
        scores=[score.model_dump() for score in payload.scores],
        class_rank=payload.class_rank,
        grade_rank=payload.grade_rank,
    )
    db.add(student)
    db.commit()
    db.refresh(student)
    return _serialize_student(student)


@router.put("/{student_id}", response_model=schemas.Student)
def update_student(
    student_id: int,
    payload: schemas.StudentUpdate,
    db: Session = Depends(get_db),
    current_user: models.Member = Depends(get_active_member),
):
    student = db.get(models.Student, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")
    if payload.student_no and payload.student_no != student.student_no:
        if db.query(models.Student).filter_by(student_no=payload.student_no).first():
            raise HTTPException(status_code=400, detail="学号已存在")
        student.student_no = payload.student_no
    for field in ["name", "class_name", "grade_name", "exam_name", "gender", "notes"]:
        value = getattr(payload, field)
        if value is not None:
            setattr(student, field, value)
    if payload.scores is not None:
        student.scores = [score.model_dump() for score in payload.scores]
    if payload.class_rank is not None:
        student.class_rank = payload.class_rank
    if payload.grade_rank is not None:
        student.grade_rank = payload.grade_rank
    db.add(student)
    db.commit()
    db.refresh(student)
    return _serialize_student(student)


@router.delete("/{student_id}", status_code=204)
def delete_student(
    student_id: int,
    db: Session = Depends(get_db),
    current_user: models.Member = Depends(get_active_member),
):
    student = db.get(models.Student, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")
    db.delete(student)
    db.commit()
    return None


def _workbook_response(wb: Workbook, filename: str) -> StreamingResponse:
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/import")
async def import_students(
    file: UploadFile = File(...),
    class_name: Optional[str] = Form(None),
    grade_name: Optional[str] = Form(None),
    exam_name: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: models.Member = Depends(get_active_member),
):
    contents = await file.read()
    try:
        wb = load_workbook(BytesIO(contents))
    except Exception as exc:  # pragma: no cover - openpyxl internal
        raise HTTPException(status_code=400, detail="无法读取 Excel 文件") from exc

    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(status_code=400, detail="表格为空")
    header_map: Dict[str, int] = {str(value).strip(): idx for idx, value in enumerate(header_row) if value}
    for required in ["姓名", "学号"]:
        if required not in header_map:
            raise HTTPException(status_code=400, detail=f"缺少必要列：{required}")

    imported = 0
    updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        name = _get_cell_value(row, header_map, "姓名")
        student_no = _get_cell_value(row, header_map, "学号")
        if not name or not student_no:
            continue
        row_class = _get_cell_value(row, header_map, "班级") or class_name
        row_grade = _get_cell_value(row, header_map, "年级") or grade_name
        row_exam = _get_cell_value(row, header_map, "考试名称") or exam_name
        row_exam_date_raw = _get_cell_value(row, header_map, "考试日期")
        row_exam_date = None
        if row_exam_date_raw:
            try:
                if isinstance(row_exam_date_raw, datetime):
                    row_exam_date = row_exam_date_raw
                elif isinstance(row_exam_date_raw, str):
                    # 尝试解析日期字符串，支持常见格式
                    for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日", "%Y.%m.%d"]:
                        try:
                            row_exam_date = datetime.strptime(row_exam_date_raw, fmt)
                            break
                        except ValueError:
                            continue
            except (ValueError, TypeError):
                pass  # 如果解析失败，保持为 None
        row_notes = _get_cell_value(row, header_map, "备注")
        scores = []
        for subject in DEFAULT_SUBJECTS:
            value = _get_cell_value(row, header_map, subject)
            if value is None or value == "":
                continue
            try:
                score_value = float(value)
            except (TypeError, ValueError):
                continue
            scores.append({"subject": subject, "score": score_value})

        existing = (
            db.query(models.Student)
                .filter(
                    models.Student.student_no == student_no,
                    models.Student.exam_name == row_exam,
                )
                .first()
        )
        if existing:
            existing.name = name
            existing.class_name = row_class
            existing.grade_name = row_grade
            existing.exam_name = row_exam
            existing.exam_date = row_exam_date
            existing.notes = row_notes
            existing.scores = scores
            db.add(existing)
            updated += 1
        else:
            student = models.Student(
                name=name,
                student_no=student_no,
                class_name=row_class,
                grade_name=row_grade,
                exam_name=row_exam,
                exam_date=row_exam_date,
                notes=row_notes,
                scores=scores,
            )
            db.add(student)
            imported += 1
    db.commit()
    return {"imported": imported, "updated": updated}


@router.post("/import-ranks")
async def import_ranks(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """导入年级排名和班级排名数据"""
    contents = await file.read()
    try:
        wb = load_workbook(BytesIO(contents))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="无法读取 Excel 文件") from exc

    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(status_code=400, detail="表格为空")

    header_map: Dict[str, int] = {str(value).strip(): idx for idx, value in enumerate(header_row) if value}

    # 验证必需列
    for required in ["姓名", "学号", "考试名称"]:
        if required not in header_map:
            raise HTTPException(status_code=400, detail=f"缺少必要列：{required}")

    updated = 0
    not_found = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None:
            continue

        name = _get_cell_value(row, header_map, "姓名")
        student_no = _get_cell_value(row, header_map, "学号")
        exam_name = _get_cell_value(row, header_map, "考试名称")

        if not student_no or not exam_name:
            continue

        # 获取排名数据
        grade_rank_value = _get_cell_value(row, header_map, "年级排名")
        class_rank_value = _get_cell_value(row, header_map, "班级排名")

        grade_rank = None
        class_rank = None

        if grade_rank_value:
            try:
                grade_rank = int(grade_rank_value)
            except (ValueError, TypeError):
                errors.append(f"第{row_idx}行：年级排名格式错误")
                continue

        if class_rank_value:
            try:
                class_rank = int(class_rank_value)
            except (ValueError, TypeError):
                errors.append(f"第{row_idx}行：班级排名格式错误")
                continue

        # 查找匹配的学生记录（根据学号和考试名称）
        existing = db.query(models.Student).filter(
            models.Student.student_no == student_no,
            models.Student.exam_name == exam_name
        ).first()

        if existing:
            # 更新排名数据
            if grade_rank is not None:
                existing.grade_rank = grade_rank
            if class_rank is not None:
                existing.class_rank = class_rank
            db.add(existing)
            updated += 1
        else:
            not_found += 1
            errors.append(f"第{row_idx}行：未找到学号 {student_no} 在考试 {exam_name} 的记录")

    db.commit()

    return {
        "updated": updated,
        "not_found": not_found,
        "errors": errors[:10] if errors else []  # 只返回前10个错误
    }


@router.get("/export")
def export_students(
    class_name: Optional[str] = Query(None),
    exam_name: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    query = db.query(models.Student)
    if class_name:
        query = query.filter(models.Student.class_name == class_name)
    if exam_name:
        query = query.filter(models.Student.exam_name == exam_name)
    students = query.order_by(models.Student.name).all()
    wb = Workbook()
    ws = wb.active
    headers = ["姓名", "学号", "班级", "年级", "考试名称", "考试日期", "备注", "总分", "均分"] + DEFAULT_SUBJECTS
    ws.append(headers)
    for student in students:
        score_map = {item.get("subject"): item.get("score") for item in (student.scores or [])}
        scores = [score_map.get(subject, "") for subject in DEFAULT_SUBJECTS]
        # 只计算 DEFAULT_SUBJECTS 中有分数的科目
        score_values = [score_map.get(subject) for subject in DEFAULT_SUBJECTS if score_map.get(subject) is not None and score_map.get(subject) != ""]
        score_values = [float(s) for s in score_values if isinstance(s, (int, float))]
        total = sum(score_values)
        avg = total / len(score_values) if score_values else 0
        exam_date_str = student.exam_date.strftime("%Y-%m-%d") if student.exam_date else ""
        ws.append(
            [
                student.name,
                student.student_no,
                student.class_name,
                student.grade_name,
                student.exam_name,
                exam_date_str,
                student.notes,
                round(total, 2),
                round(avg, 2),
            ]
            + scores
        )
    filename = "grade_export.xlsx" if not exam_name else f"grades_{exam_name}.xlsx"
    return _workbook_response(wb, filename)


def _get_cell_value(row: Any, header_map: Dict[str, int], title: str):
    idx = header_map.get(title)
    if idx is None:
        return None
    return row[idx]


def _ensure_subject_range(db: Session, subject: str) -> models.SubjectRange:
    record = db.query(models.SubjectRange).filter(models.SubjectRange.subject == subject).first()
    if not record:
        record = models.SubjectRange(subject=subject, config=[dict(item) for item in DEFAULT_RANGE_CONFIG])
        db.add(record)
        db.commit()
        db.refresh(record)
    return record


@router.get("/ranges", response_model=Dict[str, List[schemas.RangeSegment]])
def get_all_ranges(db: Session = Depends(get_db)):
    configs: Dict[str, List[schemas.RangeSegment]] = {}
    for subject in ALL_SUBJECTS:
        record = _ensure_subject_range(db, subject)
        configs[subject] = [schemas.RangeSegment(**segment) for segment in record.config]
    return configs


@router.put("/ranges/{subject}", response_model=List[schemas.RangeSegment])
def update_range(subject: str, payload: List[schemas.RangeSegment], db: Session = Depends(get_db)):
    if subject not in ALL_SUBJECTS:
        raise HTTPException(status_code=400, detail="科目不在允许范围内")
    if not payload:
        raise HTTPException(status_code=400, detail="请选择至少一个区间")
    validated = []
    for segment in payload:
        if segment.max is not None and segment.max <= segment.min:
            raise HTTPException(status_code=400, detail=f"{segment.name} 的最大值需大于最小值")
        validated.append(segment.model_dump())
    record = _ensure_subject_range(db, subject)
    record.config = validated
    db.add(record)
    db.commit()
    db.refresh(record)
    return [schemas.RangeSegment(**segment) for segment in record.config]


@router.post("/calculate-ranks")
def calculate_ranks(db: Session = Depends(get_db)):
    """自动计算所有考试的年级排名和班级排名"""
    # 获取所有考试名称和年级
    exam_grades = db.query(
        models.Student.exam_name,
        models.Student.grade_name
    ).distinct().all()

    updated = 0

    for exam_name, grade_name in exam_grades:
        if not exam_name or not grade_name:
            continue

        # 获取该年级该考试的所有学生，按总分降序排列
        students = db.query(models.Student).filter(
            models.Student.exam_name == exam_name,
            models.Student.grade_name == grade_name
        ).order_by(models.Student.total_score.desc().nullslast()).all()

        # 计算年级排名
        for rank, student in enumerate(students, 1):
            if student.total_score is not None:
                student.grade_rank = rank
                updated += 1

        # 按班级分组计算班级排名
        class_students = {}
        for student in students:
            if student.class_name:
                if student.class_name not in class_students:
                    class_students[student.class_name] = []
                class_students[student.class_name].append(student)

        # 为每个班级计算排名
        for class_name, class_student_list in class_students.items():
            # 按总分降序排列
            class_student_list.sort(key=lambda s: s.total_score if s.total_score is not None else -1, reverse=True)
            for rank, student in enumerate(class_student_list, 1):
                if student.total_score is not None:
                    student.class_rank = rank

    db.commit()
    return {"updated": updated}


@router.get("/backup")
def backup_data(db: Session = Depends(get_db)):
    """备份所有成绩数据"""
    students = db.query(models.Student).all()
    ranges = db.query(models.SubjectRange).all()

    data = {
        "students": [_serialize_student(s) for s in students],
        "ranges": {r.subject: r.config for r in ranges}
    }
    return data


@router.post("/restore")
def restore_data(data: Dict, db: Session = Depends(get_db)):
    """从备份恢复数据"""
    restored = 0

    # 恢复学生数据
    if "students" in data:
        for student_data in data["students"]:
            # 查找是否存在相同学号和考试的记录
            existing = db.query(models.Student).filter(
                models.Student.student_no == student_data.get("student_no"),
                models.Student.exam_name == student_data.get("exam_name")
            ).first()

            if existing:
                # 更新现有记录
                for key, value in student_data.items():
                    if key not in ["id", "created_at", "updated_at"]:
                        setattr(existing, key, value)
            else:
                # 创建新记录
                new_student = models.Student(**{
                    k: v for k, v in student_data.items()
                    if k not in ["id", "created_at", "updated_at"]
                })
                db.add(new_student)
            restored += 1

    # 恢复区间设置
    if "ranges" in data:
        for subject, config in data["ranges"].items():
            record = db.query(models.SubjectRange).filter(
                models.SubjectRange.subject == subject
            ).first()
            if record:
                record.config = config
            else:
                record = models.SubjectRange(subject=subject, config=config)
                db.add(record)

    db.commit()
    return {"restored": restored}


@router.delete("/clear-all")
def clear_all_data(db: Session = Depends(get_db)):
    """清空所有成绩数据（危险操作）"""
    deleted = db.query(models.Student).delete()
    db.commit()
    return {"deleted": deleted}
